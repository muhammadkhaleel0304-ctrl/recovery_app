import streamlit as st
import qrcode
from io import BytesIO
import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import plotly.express as px
from fpdf import FPDF
from io import BytesIO
import qrcode
# ---------- USERS ----------
USERS = {
    "Khaleel": "1.2341",
       "user": "1111"
}

# ---------- SESSION ----------
if "login" not in st.session_state:
    st.session_state.login = False

# ---------- LOGIN PAGE ----------
if not st.session_state.login:

    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg,#0f2027,#203a43,#2c5364);
    }

    h2, label {
        color: white !important;
        text-align: center;
    }

    .stButton>button {
        background: #00c6ff;
        color: white;
        border-radius: 10px;
        height: 40px;
        font-weight: bold;
    }

    .stButton>button:hover {
        background: #0072ff;
    }
    </style>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1,2,1])

    with col2:
        st.markdown("## 🔐 Login")

        user = st.text_input("Username")
        pwd = st.text_input("Password", type="password")

        login_btn = st.button("Login", use_container_width=True)

        if login_btn:
            if USERS.get(user) == pwd:
                st.session_state.login = True
                st.success("Login successful ✔")
                st.experimental_rerun()
            else:
                st.error("❌ Invalid username or password")

    st.stop()
# ====== یہاں اپنا 12 ڈیجٹ فکس کر دیں ======
FIXED_FIRST_12 = "421011234567"  # <-- اپنا 12 ڈیجٹ یہاں لکھیں
# ===========================================

# Custom Page Config & CSS Design
st.set_page_config(page_title="CNIC QR Generator", page_icon="🪪", layout="centered")

st.markdown(
    """
    <style>
        .header-card {
            background: linear-gradient(135deg, #1e293b, #0f172a);
            padding: 25px;
            border-radius: 16px;
            border: 1px solid #334155;
            text-align: center;
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.2);
            margin-bottom: 25px;
        }
        .header-title {
            color: #ffffff;
            font-size: 26px;
            font-weight: 700;
            margin-bottom: 6px;
        }
        .header-subtitle {
            color: #94a3b8;
            font-size: 14px;
        }
    </style>
""",
    unsafe_allow_html=True,
)

# Header Section - یہاں سے prefix ہٹا دیا
st.markdown(
    """
    <div class="header-card">
        <div class="header-title">🪪 CNIC QR Generator</div>
        <div class="header-subtitle"></div>
    </div>
""",
    unsafe_allow_html=True,
)

# Input Section with Columns Layout
col1, col2, col3 = st.columns([1, 3, 1])

with col2:
    # یوزر سے 13 ڈیجٹ لیں گے
    user_13_digit = st.text_input(
        "Just Enter CNIC NO",
        placeholder="مثال: 3720388692193",
        max_chars=13,
    )

    generate_btn = st.button(
        "✨ QR Code بنائیں", type="primary", use_container_width=True
    )

    if generate_btn:
        if user_13_digit:
            # ---------------- EXACT ORIGINAL LOGIC ----------------
            clean_13 = user_13_digit.replace("-", "").strip()

            full_cnic = FIXED_FIRST_12 + clean_13  # 12 + 13 = 25 ڈیجٹ

            if clean_13.isdigit() and len(full_cnic) == 25:
                data = full_cnic

                # FORCE FULL DATA ENCODING (Original Logic)
                qr = qrcode.QRCode(
                    version=None,  # AUTO size
                    error_correction=qrcode.constants.ERROR_CORRECT_H,
                    box_size=5,
                    border=4,
                )

                qr.add_data(data)
                qr.make(fit=True)

                img = qr.make_image(fill_color="black", back_color="white")

                buf = BytesIO()
                img.save(buf, format="PNG")

                img_bytes = buf.getvalue()
                # ------------------------------------------------------

                st.markdown("---")

                # یہاں سے "مکمل 25 ڈیجٹ" والا success ہٹا دیا
                # اور caption سے بھی full_cnic ہٹا دیا

                st.image(
                    img_bytes,
                    caption="QR Code Ready",
                    use_column_width=True,
                )

                st.download_button(
                    "📥 QR Code ڈاؤنلوڈ کریں",
                    data=img_bytes,
                    file_name=f"cnic_qr.png",  # فائل نیم سے بھی full_cnic ہٹا دیا
                    mime="image/png",
                    use_container_width=True,
                )
            else:
                st.error(
                    "❌ Invalid! Must Enter 13 Digit "
                )
        else:
            st.warning("Please Enter CNIC NO Without Dashes")

    
import streamlit as st
import pandas as pd
from io import BytesIO

st.markdown("---")
st.subheader("📁 Merge Sanction & Branch File")

# --- File Upload ---
col1, col2 = st.columns(2)
with col1:
    merge_file = st.file_uploader("Upload Merge File (Sanction No)", type=["xlsx","xls","csv"], key="merge_file")
with col2:
    branch_file = st.file_uploader("Upload Branch File (Branch Code)", type=["xlsx","xls","csv"], key="branch_file")

# --- Placeholders ---
merge_table_placeholder = st.empty()
merge_download_placeholder = st.empty()

if merge_file and branch_file:
    try:
        df_merge = pd.read_csv(merge_file) if merge_file.name.endswith(".csv") else pd.read_excel(merge_file)
        df_branch = pd.read_csv(branch_file) if branch_file.name.endswith(".csv") else pd.read_excel(branch_file)
    except Exception as e:
        merge_table_placeholder.error(f"Error reading files: {e}")
        st.stop()

    # --- Clean column names ---
    df_merge.columns = df_merge.columns.str.strip()
    df_branch.columns = df_branch.columns.str.strip()

    # --- Check required columns ---
    if 'sanctionno' not in df_merge.columns:
        st.error("Merge File must have column 'sanctionno'")
        st.stop()
    if 'branch code' not in df_branch.columns or 'branch_name' not in df_branch.columns or 'area_name' not in df_branch.columns:
        st.error("Branch File must have columns 'branch code', 'branch_name', 'area_name'")
        st.stop()

    # --- Ensure columns are string for merge ---
    df_merge['Sanction_Prefix'] = df_merge['sanctionno'].astype(str).str[:4]
    df_branch['branch code'] = df_branch['branch code'].astype(str)

    # --- Merge logic ---
    merged_df = pd.merge(
        df_merge,
        df_branch.rename(columns={'branch code':'Sanction_Prefix'}),
        on='Sanction_Prefix',
        how='left'
    )

    # --- Add Branch Name & Area Name as 3rd and 4th column ---
    if 'branch_name' in merged_df.columns and 'area_name' in merged_df.columns:
        branch_col = merged_df.pop('branch_name')
        area_col = merged_df.pop('area_name')
        merged_df.insert(2, 'Branch Name', branch_col)
        merged_df.insert(3, 'Area Name', area_col)

    # --- Display table ---
    merge_table_placeholder.dataframe(merged_df)

    # --- Download helper ---
    def to_excel(df):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Merged_Report')
        return output.getvalue()

    excel_data = to_excel(merged_df)

    # --- Download button ---
    merge_download_placeholder.download_button(
        label="📥 Download Merged File",
        data=excel_data,
        file_name="Merged_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="merge_download"
    )

else:
    merge_table_placeholder.info("Upload both Merge File and Branch File to generate merged report.")
# Upload Recovery File
uploaded_file = st.file_uploader("📁 Upload Recovery File (Excel)", type=["xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)
    df['recovery_date'] = pd.to_datetime(df['recovery_date'], errors='coerce')
    df.dropna(subset=['recovery_date'], inplace=True)
    df['day'] = df['recovery_date'].dt.day

    def get_range(day):
        if 1 <= day <= 10:
            return "1-10"
        elif 11 <= day <= 20:
            return "11-20"
        elif 21 <= day <= 31:
            return "21-31"
        return "Unknown"

    df['range'] = df['day'].apply(get_range)

    st.write("### 📄 Complete Recovery Data")
    st.dataframe(df)

    # Summary
    summary = df.groupby(['branch_id', 'range']).agg({
        'amount': 'sum',
        'receipt_no': 'count'
    }).reset_index()

    branch_totals = df.groupby('branch_id')['amount'].sum().reset_index().rename(columns={'amount': 'total_amount'})
    summary = summary.merge(branch_totals, on='branch_id')
    summary['percentage'] = (summary['amount'] / summary['total_amount']) * 100

    st.subheader("📊 Branch-wise Recovery Summary")
    st.dataframe(summary.style.format({
        'amount': 'Rs {:,.0f}',
        'percentage': '{:.2f}%'
    }))

    # Chart
    st.subheader("📈 Recovery Chart by Date Range")
    fig = px.bar(summary, x='branch_id', y='amount', color='range',
                 barmode='group',
                 text=summary['percentage'].apply(lambda x: f"{x:.1f}%"),
                 labels={'amount': 'Amount Recovered', 'branch_id': 'Branch'})
    fig.update_traces(textposition='outside')
    fig.update_layout(xaxis_title="Branch", yaxis_title="Amount", legend_title="Date Range")
    st.plotly_chart(fig, use_container_width=True)

    # Pivot Table
    st.subheader("📌 Pivot Table (Branch → Project → Date)")
    pivot_df = df.groupby(['branch_id', 'project', 'recovery_date']).agg(
        Receipts=('receipt_no', 'count'),
        Amount=('amount', 'sum')
    ).reset_index()

    st.dataframe(pivot_df)

    # PDF Class
    class PDF(FPDF):
        def header(self):
            pass
        def footer(self):
            pass

    # Branch-wise PDF downloads
    st.subheader("📥 Download Branch-wise Pivot Table PDFs")
    for branch, branch_df in pivot_df.groupby('branch_id'):
        branch_pdf = PDF()
        branch_pdf.set_auto_page_break(auto=True, margin=15)
        branch_pdf.add_page()
        branch_pdf.set_font("Arial", 'B', 14)
        branch_pdf.cell(0, 10, f"Branch: {branch}", ln=True, align='C')

        branch_total_amount = 0
        branch_total_receipts = 0

        for project, proj_df in branch_df.groupby('project'):
            branch_pdf.set_font("Arial", 'B', 12)
            branch_pdf.cell(0, 8, f"Project: {project}", ln=True)

            # Table Header
            branch_pdf.set_font("Arial", 'B', 10)
            branch_pdf.cell(40, 8, "Date", border=1, align='C')
            branch_pdf.cell(40, 8, "Receipts", border=1, align='C')
            branch_pdf.cell(40, 8, "Amount", border=1, align='C')
            branch_pdf.ln()

            project_total_amount = 0
            project_total_receipts = 0

            branch_pdf.set_font("Arial", '', 10)
            for _, row in proj_df.iterrows():
                date_str = row['recovery_date'].strftime('%Y-%m-%d') if pd.notnull(row['recovery_date']) else ''
                branch_pdf.cell(40, 8, date_str, border=1)
                branch_pdf.cell(40, 8, str(row['Receipts']), border=1, align='C')
                branch_pdf.cell(40, 8, f"Rs {row['Amount']:,.0f}", border=1, align='R')
                branch_pdf.ln()
                project_total_receipts += row['Receipts']
                project_total_amount += row['Amount']

            # Project total
            branch_pdf.set_font("Arial", 'B', 10)
            branch_pdf.cell(40, 8, "Project Total", border=1)
            branch_pdf.cell(40, 8, str(project_total_receipts), border=1, align='C')
            branch_pdf.cell(40, 8, f"Rs {project_total_amount:,.0f}", border=1, align='R')
            branch_pdf.ln(10)

            branch_total_receipts += project_total_receipts
            branch_total_amount += project_total_amount

        # Branch total
        branch_pdf.set_font("Arial", 'B', 11)
        branch_pdf.cell(40, 8, "Branch Total", border=1)
        branch_pdf.cell(40, 8, str(branch_total_receipts), border=1, align='C')
        branch_pdf.cell(40, 8, f"Rs {branch_total_amount:,.0f}", border=1, align='R')

        pdf_bytes = branch_pdf.output(dest='S').encode('latin1')
        st.download_button(
            label=f"📥 Download PDF for Branch {branch}",
            data=pdf_bytes,
            file_name=f"Branch_{branch}.pdf",
            mime="application/pdf"
        )
st.subheader("📥 Upload Due List and Recovery File for Overdue Detection")

dolist_file = st.file_uploader("📄 Due List Upload", type=["xlsx"], key="dolist")
recovery_file2 = st.file_uploader("📄 Recovery File Upload", type=["xlsx"], key="recovery2")

if dolist_file and recovery_file2:
    dolist_df = pd.read_excel(dolist_file)
    recovery_df2 = pd.read_excel(recovery_file2)

    dolist_df['Sanction No'] = dolist_df['Sanction No'].astype(str).str.strip()
    recovery_df2['Sanction No'] = recovery_df2['Sanction No'].astype(str).str.strip()

    overdue_df = dolist_df[~dolist_df['Sanction No'].isin(recovery_df2['Sanction No'])]
    st.subheader("❗ Overdue List")
    st.write(f"🔢 Total Overdue: {len(overdue_df)}")
    st.dataframe(overdue_df)

# Final Overdue via Terabyte
st.subheader("📥 Upload Terabyte File (Final Overdue)")

terabyte_file = st.file_uploader("📄 Terabyte File Upload", type=["xlsx"], key="terabyte")

if terabyte_file and 'overdue_df' in locals() and not overdue_df.empty:
    terabyte_df = pd.read_excel(terabyte_file)
    terabyte_df['Sanction No'] = terabyte_df['Sanction No'].astype(str).str.strip()
    overdue_df['Sanction No'] = overdue_df['Sanction No'].astype(str).str.strip()

    final_overdue_df = overdue_df[~overdue_df['Sanction No'].isin(terabyte_df['Sanction No'])]

    st.subheader("🚨 Final Overdue Cases")
    st.write(f"🔢 Total Final Overdue: {len(final_overdue_df)}")
    st.dataframe(final_overdue_df)

    # Full PDF: Branch-wise + Date-wise
    full_pdf = FPDF()
    full_pdf.set_auto_page_break(auto=True, margin=15)

    if 'branch_id' not in final_overdue_df.columns:
        final_overdue_df['branch_id'] = 'Unknown'

    for branch in final_overdue_df['branch_id'].unique():
        data = final_overdue_df[final_overdue_df['branch_id'] == branch]
        full_pdf.add_page()
        full_pdf.set_font("Arial", 'B', 12)
        full_pdf.cell(200, 10, txt=f"Branch: {branch}", ln=True, align='C')

        full_pdf.set_font("Arial", size=10)
        full_pdf.cell(10, 10, "Sr#", 1)
        full_pdf.cell(70, 10, "Name", 1)
        full_pdf.cell(60, 10, "Sanction No", 1)
        full_pdf.ln()

        for i, (_, row) in enumerate(data.iterrows(), start=1):
            full_pdf.cell(10, 10, str(i), 1)
            full_pdf.cell(70, 10, str(row.get('Name', '')), 1)
            full_pdf.cell(60, 10, str(row.get('Sanction No', '')), 1)
            full_pdf.ln()

    full_pdf_output = full_pdf.output(dest='S').encode('latin1')
    st.download_button("📥 Download Final Overdue PDF (Branch-wise)", full_pdf_output, "final_overdue.pdf", "application/pdf")
# 🔽 Separate Branch-wise PDF Downloads
    st.subheader("📂 Download Final Overdue Branch-wise PDFs")

    branch_pdfs = {}

    for branch in final_overdue_df['branch_id'].unique():
        branch_data = final_overdue_df[final_overdue_df['branch_id'] == branch]

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(200, 10, txt=f"Branch: {branch}", ln=True, align='C')

        pdf.set_font("Arial", size=10)
        pdf.cell(10, 10, "Sr#", 1)
        pdf.cell(70, 10, "Name", 1)
        pdf.cell(60, 10, "Sanction No", 1)
        pdf.ln()

        for i, (_, row) in enumerate(branch_data.iterrows(), start=1):
            pdf.cell(10, 10, str(i), 1)
            pdf.cell(70, 10, str(row.get('Name', '')), 1)
            pdf.cell(60, 10, str(row.get('Sanction No', '')), 1)
            pdf.ln()

        pdf_bytes = pdf.output(dest='S').encode('latin1')
        branch_pdfs[branch] = pdf_bytes

    for branch, pdf_data in branch_pdfs.items():
        st.download_button(
            label=f"📥 Download PDF for Branch: {branch}",
            data=pdf_data,
            file_name=f"final_overdue_branch_{branch}.pdf",
            mime="application/pdf"
        )
import streamlit as st
import pandas as pd
import os
import io
import zipfile
from datetime import datetime
from dateutil.relativedelta import relativedelta
from fpdf import FPDF
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
import plotly.express as px

st.title("🏦 Recovery & Reports App")

# --------------------
# File upload widgets (unique keys)
# --------------------
do_file = st.file_uploader("Upload Do List", type=["xlsx", "xls"], key="uploader_do")
recovery_file = st.file_uploader("Upload Recovery File", type=["xlsx", "xls"], key="uploader_recovery")
terabyte_file = st.file_uploader("Upload Terabyte File (Optional)", type=["xlsx", "xls"], key="uploader_terabyte")

# --------------------
# When Do + Recovery uploaded -> main logic
# --------------------
if do_file and recovery_file:
    # Read files
    do_df = pd.read_excel(do_file)
    recovery_df = pd.read_excel(recovery_file)

    # Normalize column names
    for df in [do_df, recovery_df]:
        df.columns = df.columns.str.strip()

    # --------------------
    # Overdue logic (Final Overdue List)
    # --------------------
    if 'Sanction No' not in do_df.columns or 'Sanction No' not in recovery_df.columns:
        st.error("Both Do List and Recovery File must contain 'Sanction No' column.")
    else:
        overdue_df = do_df[~do_df['Sanction No'].astype(str).str.strip().isin(
            recovery_df['Sanction No'].astype(str).str.strip()
        )].copy()

        st.subheader("🕒 Final Overdue List")
        st.dataframe(overdue_df)

        # Branch-wise Overdue PDF (as ZIP)
        if not overdue_df.empty:
            st.subheader("📁 Download Branch-wise Final Overdue (ZIP)")
            overdue_df.columns = overdue_df.columns.str.strip()
            branches = overdue_df['branch_id'].astype(str).unique()

            zip_buf_overdue = io.BytesIO()
            with zipfile.ZipFile(zip_buf_overdue, "a", zipfile.ZIP_DEFLATED) as zf:
                for branch in branches:
                    branch_data = overdue_df[overdue_df['branch_id'].astype(str) == str(branch)]

                    pdf = FPDF()
                    pdf.add_page()
                    pdf.set_font("Arial", size=10)
                    pdf.cell(0, 10, f"Branch: {branch}", ln=True)

                    # Header
                    pdf.set_font("Arial", "B", 10)
                    pdf.cell(10, 10, "Sr#", 1)
                    pdf.cell(60, 10, "Name", 1)
                    pdf.cell(50, 10, "Sanction No", 1)
                    # include mobile if present
                    if "Mobile No" in branch_data.columns:
                        pdf.cell(50, 10, "Mobile No", 1)
                    pdf.ln()

                    pdf.set_font("Arial", size=9)
                    for i, (_, row) in enumerate(branch_data.iterrows(), start=1):
                        pdf.cell(10, 10, str(i), 1)
                        pdf.cell(60, 10, str(row.get("Name", ""))[:25], 1)
                        pdf.cell(50, 10, str(row.get("Sanction No", "")), 1)
                        if "Mobile No" in branch_data.columns:
                            pdf.cell(50, 10, str(row.get("Mobile No", "")), 1)
                        pdf.ln()

                    pdf_bytes = pdf.output(dest="S").encode("latin1")
                    zf.writestr(f"{branch}_Final_Overdue.pdf", pdf_bytes)

            zip_buf_overdue.seek(0)
            st.download_button(
                label="⬇️ Download Branch-wise Overdue ZIP",
                data=zip_buf_overdue.getvalue(),
                file_name="Final_Overdue_BranchWise.zip",
                mime="application/zip",
                key="download_overdue_zip"
            )
        else:
            st.info("No overdue records found.")

    # --------------------
    # Recovery this month / matched recoveries and summary / charts (original)
    # --------------------
    # Sanction No normalization
    do_df['Sanction No'] = do_df['Sanction No'].astype(str).str.strip()
    recovery_df['Sanction No'] = recovery_df['Sanction No'].astype(str).str.strip()

    # Parse recovery date if present
    if 'recovery_date' in recovery_df.columns:
        recovery_df['recovery_date'] = pd.to_datetime(recovery_df['recovery_date'], errors='coerce')
    else:
        # if no recovery_date column, try common names or create empty
        recovery_df['recovery_date'] = pd.NaT

    current_month = pd.Timestamp.now().month
    current_year = pd.Timestamp.now().year

    recovery_this_month = recovery_df[
        (recovery_df['recovery_date'].dt.month == current_month) &
        (recovery_df['recovery_date'].dt.year == current_year)
    ] if not recovery_df['recovery_date'].isna().all() else recovery_df.iloc[0:0]

    recovered = recovery_this_month[recovery_this_month['Sanction No'].isin(do_df['Sanction No'])]

    # Due summary per branch
    if 'branch_id' not in do_df.columns:
        do_df['branch_id'] = do_df.get('Branch', '')
    if 'branch_id' not in recovery_df.columns:
        recovery_df['branch_id'] = recovery_df.get('Branch Code', '')

    due_summary = do_df.groupby('branch_id')['Sanction No'].count().reset_index()
    due_summary.columns = ['branch_id', 'total_due']

    recovered_summary = recovered.groupby('branch_id')['Sanction No'].count().reset_index()
    recovered_summary.columns = ['branch_id', 'recovered']

    summary = due_summary.merge(recovered_summary, on='branch_id', how='left').fillna(0)
    summary['remaining'] = summary['total_due'] - summary['recovered']
    summary['recovery_percent'] = (summary['recovered'] / summary['total_due'].replace(0, pd.NA)) * 100
    summary['recovery_percent'] = summary['recovery_percent'].fillna(0)

    st.subheader("📋 Branch-wise Recovery Summary (This Month)")
    st.dataframe(summary.style.format({
        'total_due': '{:,.0f}',
        'recovered': '{:,.0f}',
        'remaining': '{:,.0f}',
        'recovery_percent': '{:.2f} %'
    }))

    # Chart: recovery percent by branch
    try:
        fig = px.bar(
            summary,
            x='branch_id',
            y='recovery_percent',
            text=summary['recovery_percent'].apply(lambda x: f"{x:.1f}%"),
            labels={'branch_id': 'Branch', 'recovery_percent': 'Recovery %'},
            title='📈 Recovery % by Branch (This Month)'
        )
        fig.update_traces(textposition='outside')
        st.plotly_chart(fig, use_container_width=True)
    except Exception:
        pass

    # Debugging info
    st.subheader("🛠 Debugging Info")
    st.write("Total Due List Entries:", len(do_df))
    st.write("Recovery Entries This Month:", len(recovery_this_month))
    st.write("Matched Recoveries:", len(recovered))

# --------------------
# TERABYTE SECTION (kept as originally present in your code)
# --------------------
# This block expects a terabyte upload (can be optional)
terabyte_pdf_file = st.file_uploader("Upload Terabyte Excel (for branch receipts PDF)", type=["xls", "xlsx"], key="uploader_terabyte_pdf")

if terabyte_pdf_file is not None:
    df_tera = pd.read_excel(terabyte_pdf_file)

    required_cols = ["Sanction No", "Recovery Date", "Receipt No", "Credit Amount", "Branch Code"]
    missing = [c for c in required_cols if c not in df_tera.columns]
    if missing:
        st.error(f"Uploaded Terabyte file must contain columns: {', '.join(missing)}")
    else:
        df_tera["Recovery Date"] = pd.to_datetime(df_tera["Recovery Date"], errors='coerce').dt.date
        df_tera.insert(0, "Serial No", range(1, len(df_tera) + 1))
        df_tera = df_tera[["Serial No", "Sanction No", "Recovery Date", "Receipt No", "Credit Amount", "Branch Code"]]

        st.write("### Terabyte Branch-wise Preview")
        st.dataframe(df_tera.head())

        # Branch-wise downloadable PDFs using reportlab (kept original approach)
        branches = df_tera["Branch Code"].unique()
        for branch in branches:
            branch_df = df_tera[df_tera["Branch Code"] == branch]

            st.write(f"#### Branch {branch} Summary")
            st.dataframe(branch_df)

            # Create in-memory PDF and show download button per branch
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            elements.append(Paragraph(f"Branch Code: {branch}", styles['Heading1']))
            elements.append(Spacer(1, 12))

            table_data = [list(branch_df.columns)] + branch_df.values.tolist()
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))

            # Branch summary
            branch_summary = pd.DataFrame({
                "Branch Code": [branch],
                "Total Receipts": [len(branch_df)],
                "Total Amount": [branch_df["Credit Amount"].sum()]
            })
            elements.append(Paragraph("Branch Summary", styles['Heading2']))
            branch_table = Table([list(branch_summary.columns)] + branch_summary.values.tolist())
            branch_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(branch_table)
            elements.append(Spacer(1, 20))

            # Date-wise summary
            date_summary = branch_df.groupby("Recovery Date").agg(
                Receipts_Count=("Receipt No", "count"),
                Amount_Sum=("Credit amount", "sum")
            ).reset_index()
            elements.append(Paragraph("Date-wise Summary", styles['Heading2']))
            date_table = Table([list(date_summary.columns)] + date_summary.values.tolist())
            date_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(date_table)

            doc.build(elements)
            buf.seek(0)

            st.download_button(
                label=f"Download Branch {branch} PDF",
                data=buf.getvalue(),
                file_name=f"branch_{branch}.pdf",
                mime="application/pdf",
                key=f"download_terabyte_{branch}"
            )

# --------------------
# NEW: Branch-wise Recovery PDFs from uploaded Recovery File
# (This is the single addition you asked for — everything else left intact)
# --------------------
# This button will be available if user uploaded a recovery_file earlier.
if 'recovery_file' in locals() or recovery_file is not None:
    # Use the uploaded recovery_file object if present
    # (We attempt to read it again safely here)
    try:
        if recovery_file is not None:
            rec_df_for_pdf = pd.read_excel(recovery_file)
        else:
            rec_df_for_pdf = None
    except Exception:
        rec_df_for_pdf = None

    if rec_df_for_pdf is not None and not rec_df_for_pdf.empty:
        st.subheader("📄 Generate Branch-wise PDFs from Recovery File")
        st.write("This will create a PDF per branch (recovery_date,amount, Name, Sanction No)")

        if st.button("⬇️ Generate Branch-wise Recovery PDFs", key="gen_recovery_pdfs_btn"):
            # Normalize and ensure columns
            rec_df = rec_df_for_pdf.copy()
            rec_df.columns = rec_df.columns.str.strip()

            # Ensure these columns exist or create empty
            for col in ["branch_id", "recovery_date", "amount", "Name", "Sanction No"]:
                if col not in rec_df.columns:
                    rec_df[col] = ""

            # Format Date column
            try:
                rec_df["Date"] = pd.to_datetime(rec_df["recovery_date"], errors='coerce').dt.strftime("%d-%m-%y")
            except Exception:
                rec_df["recovery_date"] = rec_df["recovery_date"].astype(str)

            branches = rec_df["branch_id"].astype(str).unique()
            zip_buffer = io.BytesIO()

            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED) as zf:
                for branch in branches:
                    branch_data = rec_df[rec_df["branch_id"].astype(str) == str(branch)]

                    pdf = FPDF()
                    pdf.add_page()
                    pdf.set_font("Arial", "B", 14)
                    pdf.cell(0, 10, f"Branch: {branch}", ln=True, align="L")
                    pdf.ln(6)

                    # Header
                    pdf.set_font("Arial", "B", 10)
                    pdf.cell(12, 8, "Sr#", 1, 0, "C")
                    pdf.cell(50, 8, "Date", 1, 0, "C")
                    pdf.cell(40, 8, "amount", 1, 0, "C")
                    pdf.cell(60, 8, "Name", 1, 0, "C")
                    pdf.cell(40, 8, "Sanction No", 1, 1, "C")

                    pdf.set_font("Arial", "", 9)
                    total_amount = 0.0
                    for i, (_, row) in enumerate(branch_data.iterrows(), start=1):
                        pdf.cell(12, 8, str(i), 1, 0, "C")
                        pdf.cell(50, 8, str(row.get("Date", ""))[:10], 1, 0, "C")
                        pdf.cell(40, 8, str(row.get("amount", "")), 1, 0, "R")
                        pdf.cell(60, 8, str(row.get("Name", ""))[:25], 1, 0, "L")
                        pdf.cell(40, 8, str(row.get("Sanction No", "")), 1, 1, "C")
                        try:
                            total_amount += float(row.get("amount", 0) if row.get("amount", 0) != "" else 0)
                        except Exception:
                            pass

                    # Total row
                    pdf.set_font("Arial", "B", 10)
                    pdf.cell(62, 8, "Total", 1)
                    pdf.cell(40, 8, f"{total_amount:,.2f}", 1)
                    pdf.ln(8)

                    pdf_bytes = pdf.output(dest="S").encode("latin1")
                    zf.writestr(f"{branch}_Recovery.pdf", pdf_bytes)

            zip_buffer.seek(0)
            st.download_button(
                label="📦 Download Branch-wise Recovery PDFs (ZIP)",
                data=zip_buffer.getvalue(),
                file_name="Branch_Wise_Recovery_PDFs.zip",
                mime="application/zip",
                key="download_recovery_zip"
            )
    else:
        # No recovery file data to generate from
        pass

import streamlit as st
import pandas as pd
import re

st.header("📂 Merge CSV Files (Skip first 2 rows)")

def clean_colname(name):
    return re.sub(r'[^a-z0-9]', '', str(name).lower())

# --- Users select multiple CSV files ---
uploaded_files = st.file_uploader(
    "Upload your CSV files",
    type="csv",
    accept_multiple_files=True
)

merged_data = []
missing_sanction_files = []

if uploaded_files:
    for uploaded_file in uploaded_files:
        try:
            # Skip first 2 rows
            df = pd.read_csv(uploaded_file, skiprows=2)

            # Clean columns
            df.columns = [clean_colname(col) for col in df.columns]

            # Check for Sanction No column
            possible_names = ["sanctionno", "sanctionnumber", "sactionno"]
            sanction_col = next((col for col in df.columns if col in possible_names), None)

            if sanction_col:
                merged_data.append(df)
            else:
                missing_sanction_files.append(uploaded_file.name)

        except Exception as e:
            st.error(f"Error reading {uploaded_file.name}: {e}")

    # --- Show warning for files without Sanction No ---
    if missing_sanction_files:
        st.warning("No 'Sanction No' column found in these files:")
        for f in missing_sanction_files:
            st.write(f"- {f}")

    # --- Merge and allow download ---
    if merged_data:
        final_df = pd.concat(merged_data, ignore_index=True)
        st.success(f"Merged {len(merged_data)} CSV files! Total rows: {len(final_df)}")

        csv_download = final_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="⬇ Download Merged CSV",
            data=csv_download,
            file_name="merged_due_list.csv",
            mime="text/csv"
        )
else:
    st.info("Please upload at least one CSV file to merge.")
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# =========================================================
# RANGE-WISE RECOVERY COMPARISON
# =========================================================

st.markdown("""
<style>
.range-title {
    background: linear-gradient(135deg,#063b66,#0876b9);
    color:white;
    padding:18px;
    border-radius:12px;
    text-align:center;
    font-size:26px;
    font-weight:800;
    margin-bottom:18px;
}

.info-box {
    background:#f5f9ff;
    border:1px solid #cbdff5;
    border-radius:10px;
    padding:12px;
}
</style>
""", unsafe_allow_html=True)

st.markdown(
    '<div class="range-title">📊 RANGE-WISE RECOVERY COMPARISON REPORT</div>',
    unsafe_allow_html=True
)


# =========================================================
# EXCEL UPLOAD
# =========================================================

st.subheader("📤 Upload Recovery Excel")

uploaded_file = st.file_uploader(
    "Upload your Recovery Excel File",
    type=["xlsx", "xls"],
    key="range_recovery_upload"
)

if uploaded_file is None:

    st.info(
        "Please upload your Recovery Excel file. "
        "The report will be generated automatically after upload."
    )

    st.stop()


# =========================================================
# READ EXCEL
# =========================================================

try:

    excel_file = pd.ExcelFile(uploaded_file)

    sheet_name = st.selectbox(
        "Select Sheet",
        excel_file.sheet_names
    )

    raw_df = pd.read_excel(
        uploaded_file,
        sheet_name=sheet_name
    )

except Exception as e:

    st.error(f"Excel file read نہیں ہو سکی: {e}")
    st.stop()


# =========================================================
# CLEAN COLUMN NAMES
# =========================================================

raw_df.columns = (
    raw_df.columns
    .astype(str)
    .str.strip()
    .str.lower()
    .str.replace(" ", "_")
    .str.replace("-", "_")
)


# =========================================================
# FIND COLUMNS AUTOMATICALLY
# =========================================================

def find_column(df, possible_names):

    for col in possible_names:

        if col in df.columns:
            return col

    return None


# ---------------------------------------------------------
# AREA COLUMN
# ---------------------------------------------------------

area_col = find_column(
    raw_df,
    [
        "area",
        "area_name",
        "region",
        "zone"
    ]
)


# ---------------------------------------------------------
# BRANCH NAME COLUMN
# ---------------------------------------------------------

branch_name_col = find_column(
    raw_df,
    [
        "branch_name",
        "branchname",
        "branch_title",
        "branch"
    ]
)


# ---------------------------------------------------------
# DATE COLUMN
# ---------------------------------------------------------

date_col = find_column(
    raw_df,
    [
        "recovery_date",
        "recoverydate",
        "date",
        "recovery_date_",
        "receipt_date",
        "transaction_date"
    ]
)


# =========================================================
# MANUAL COLUMN SELECTION
# =========================================================

st.subheader("🔧 Column Mapping")

col1, col2, col3 = st.columns(3)


# ---------------------------------------------------------
# AREA
# ---------------------------------------------------------

with col1:

    if area_col is None:

        area_col = st.selectbox(
            "Area Column",
            ["None"] + list(raw_df.columns),
            key="area_mapping"
        )

        if area_col == "None":
            area_col = None

    else:

        st.success(
            f"Area: {area_col}"
        )


# ---------------------------------------------------------
# BRANCH NAME
# ---------------------------------------------------------

with col2:

    if branch_name_col is None:

        branch_name_col = st.selectbox(
            "Branch Name Column",
            ["None"] + list(raw_df.columns),
            key="branch_name_mapping"
        )

        if branch_name_col == "None":
            branch_name_col = None

    else:

        st.success(
            f"Branch Name: {branch_name_col}"
        )


# ---------------------------------------------------------
# RECOVERY DATE
# ---------------------------------------------------------

with col3:

    if date_col is None:

        date_col = st.selectbox(
            "Recovery Date Column",
            ["None"] + list(raw_df.columns),
            key="date_mapping"
        )

        if date_col == "None":
            date_col = None

    else:

        st.success(
            f"Recovery Date: {date_col}"
        )


# =========================================================
# REQUIRED COLUMN CHECK
# =========================================================

if date_col is None:

    st.error(
        "Recovery Date column نہیں ملی۔ "
        "براہ کرم اوپر سے Recovery Date والی column select کریں۔"
    )

    st.stop()


if area_col is None:

    st.error(
        "Area column نہیں ملی۔ "
        "براہ کرم Area والی column select کریں۔"
    )

    st.stop()


if branch_name_col is None:

    st.error(
        "Branch Name column نہیں ملی۔ "
        "براہ کرم Branch Name والی column select کریں۔"
    )

    st.stop()


# =========================================================
# DATE CONVERSION
# =========================================================

df = raw_df.copy()

df["_recovery_date"] = pd.to_datetime(
    df[date_col],
    errors="coerce",
    dayfirst=True
)

invalid_dates = df["_recovery_date"].isna().sum()

if invalid_dates > 0:

    st.warning(
        f"{invalid_dates} rows میں valid recovery date نہیں ملی۔ "
        "یہ rows report میں شامل نہیں ہوں گی۔"
    )

df = df[
    df["_recovery_date"].notna()
].copy()


if len(df) == 0:

    st.error(
        "کوئی valid Recovery Date نہیں ملی۔"
    )

    st.stop()


# =========================================================
# AREA INFORMATION
# =========================================================

df["_area"] = (
    df[area_col]
    .fillna("")
    .astype(str)
    .str.strip()
)


# =========================================================
# BRANCH INFORMATION
# =========================================================

df["_branch_name"] = (
    df[branch_name_col]
    .fillna("")
    .astype(str)
    .str.strip()
)


# =========================================================
# REMOVE EMPTY AREA / BRANCH
# =========================================================

df = df[
    (df["_area"] != "") &
    (df["_branch_name"] != "")
].copy()


if df.empty:

    st.error(
        "Area یا Branch Name میں کوئی valid data نہیں ملا۔"
    )

    st.stop()


# =========================================================
# MONTH
# =========================================================

df["_month"] = (
    df["_recovery_date"]
    .dt.to_period("M")
)


available_months = sorted(
    df["_month"]
    .dropna()
    .unique()
)


if len(available_months) == 0:

    st.error(
        "کوئی month data نہیں ملا۔"
    )

    st.stop()


month_names = [
    x.strftime("%b-%y")
    for x in available_months
]


# =========================================================
# MONTH SELECTION
# =========================================================

m1, m2 = st.columns(2)


with m1:

    from_month_label = st.selectbox(
        "From Month",
        month_names,
        index=max(0, len(month_names) - 2)
    )


with m2:

    to_month_label = st.selectbox(
        "To Month",
        month_names,
        index=len(month_names) - 1
    )


# =========================================================
# CONVERT MONTH LABEL TO PERIOD
# =========================================================

try:

    from_month = pd.Period(
        pd.to_datetime(
            "01-" + from_month_label,
            format="%d-%b-%y"
        ),
        freq="M"
    )

    to_month = pd.Period(
        pd.to_datetime(
            "01-" + to_month_label,
            format="%d-%b-%y"
        ),
        freq="M"
    )

except Exception as e:

    st.error(
        f"Month selection error: {e}"
    )

    st.stop()


# =========================================================
# OPTIONAL AREA FILTER
# =========================================================

area_filter_values = [
    "All"
] + sorted(
    df["_area"]
    .dropna()
    .astype(str)
    .unique()
    .tolist()
)


selected_area = st.selectbox(
    "📍 Area Filter",
    area_filter_values
)


if selected_area != "All":

    df = df[
        df["_area"] == selected_area
    ].copy()


# =========================================================
# RANGE FUNCTION
# =========================================================

def get_range(day):

    if 1 <= day <= 5:
        return "1-5"

    elif 6 <= day <= 10:
        return "6-10"

    elif 11 <= day <= 15:
        return "11-15"

    elif 16 <= day <= 25:
        return "16-25"

    else:
        return ">25"


df["_day"] = (
    df["_recovery_date"]
    .dt.day
)

df["_range"] = (
    df["_day"]
    .apply(get_range)
)


# =========================================================
# REPORT FUNCTION
# =========================================================

def create_month_report(data, month_period):

    temp = data[
        data["_month"] == month_period
    ].copy()


    ranges = [
        "1-5",
        "6-10",
        "11-15",
        "16-25",
        ">25"
    ]


    if temp.empty:

        return pd.DataFrame()


    result = []


    # -----------------------------------------------------
    # AREA + BRANCH LIST
    # -----------------------------------------------------

    branch_list = (
        temp[
            [
                "_area",
                "_branch_name"
            ]
        ]
        .drop_duplicates()
        .sort_values(
            [
                "_area",
                "_branch_name"
            ]
        )
    )


    # -----------------------------------------------------
    # LOOP AREA + BRANCH
    # -----------------------------------------------------

    for _, branch in branch_list.iterrows():

        area = branch["_area"]

        name = branch["_branch_name"]


        branch_data = temp[
            (temp["_area"] == area) &
            (temp["_branch_name"] == name)
        ]


        row = {

            "Area": area,

            "Branch Name": name

        }


        total = 0


        for r in ranges:

            count = (
                branch_data["_range"] == r
            ).sum()


            row[r] = int(count)


            total += count


        row["Total"] = int(total)


        result.append(row)


    return pd.DataFrame(result)


# =========================================================
# CREATE REPORTS
# =========================================================

month1_df = create_month_report(
    df,
    from_month
)


month2_df = create_month_report(
    df,
    to_month
)


# =========================================================
# EMPTY CHECK
# =========================================================

if month1_df.empty:

    st.warning(
        f"{from_month_label} میں کوئی recovery record نہیں ملا۔"
    )


if month2_df.empty:

    st.warning(
        f"{to_month_label} میں کوئی recovery record نہیں ملا۔"
    )


# =========================================================
# MERGE AREAS + BRANCHES
# =========================================================

all_branches = pd.concat(
    [
        month1_df[
            [
                "Area",
                "Branch Name"
            ]
        ]
        if not month1_df.empty
        else pd.DataFrame(),

        month2_df[
            [
                "Area",
                "Branch Name"
            ]
        ]
        if not month2_df.empty
        else pd.DataFrame()
    ],
    ignore_index=True
).drop_duplicates()


if all_branches.empty:

    st.error(
        "Area / Branch data نہیں ملا۔"
    )

    st.stop()


# =========================================================
# RANGES
# =========================================================

ranges = [
    "1-5",
    "6-10",
    "11-15",
    "16-25",
    ">25"
]


# =========================================================
# MONTH 1
# =========================================================

if not month1_df.empty:

    month1_df = month1_df.rename(
        columns={
            r: f"{from_month_label} | {r}"
            for r in ranges
        }
        |
        {
            "Total":
            f"{from_month_label} | Total"
        }
    )


else:

    month1_df = all_branches.copy()


    for r in ranges:

        month1_df[
            f"{from_month_label} | {r}"
        ] = 0


    month1_df[
        f"{from_month_label} | Total"
    ] = 0


# =========================================================
# MONTH 2
# =========================================================

if not month2_df.empty:

    month2_df = month2_df.rename(
        columns={
            r: f"{to_month_label} | {r}"
            for r in ranges
        }
        |
        {
            "Total":
            f"{to_month_label} | Total"
        }
    )


else:

    month2_df = all_branches.copy()


    for r in ranges:

        month2_df[
            f"{to_month_label} | {r}"
        ] = 0


    month2_df[
        f"{to_month_label} | Total"
    ] = 0


# =========================================================
# MERGE MONTHS
# =========================================================

comparison = all_branches.merge(
    month1_df,
    on=[
        "Area",
        "Branch Name"
    ],
    how="left"
)


comparison = comparison.merge(
    month2_df,
    on=[
        "Area",
        "Branch Name"
    ],
    how="left"
)


comparison = comparison.fillna(0)


# =========================================================
# DIFFERENCE COLUMNS
# =========================================================

for r in ranges:

    comparison[
        f"Diff | {r}"
    ] = (
        comparison[
            f"{to_month_label} | {r}"
        ]
        -
        comparison[
            f"{from_month_label} | {r}"
        ]
    )


# ---------------------------------------------------------
# TOTAL DIFFERENCE
# ---------------------------------------------------------

comparison[
    "Diff | Total"
] = (
    comparison[
        f"{to_month_label} | Total"
    ]
    -
    comparison[
        f"{from_month_label} | Total"
    ]
)


# ---------------------------------------------------------
# PERCENTAGE DIFFERENCE
# ---------------------------------------------------------

comparison[
    "% Diff"
] = np.where(

    comparison[
        f"{from_month_label} | Total"
    ] == 0,

    0,

    (
        comparison["Diff | Total"]
        /
        comparison[
            f"{from_month_label} | Total"
        ]
    ) * 100
)


# =========================================================
# SERIAL NUMBER
# =========================================================

comparison.insert(
    0,
    "Sr.",
    range(
        1,
        len(comparison) + 1
    )
)


# =========================================================
# GRAND TOTAL
# =========================================================

grand_total = {}


for col in comparison.columns:

    if col == "Sr.":

        grand_total[col] = ""


    elif col == "Area":

        grand_total[col] = ""


    elif col == "Branch Name":

        grand_total[col] = "GRAND TOTAL"


    elif col == "% Diff":

        old_total = comparison[
            f"{from_month_label} | Total"
        ].sum()


        new_total = comparison[
            f"{to_month_label} | Total"
        ].sum()


        grand_total[col] = (

            (
                (new_total - old_total)
                /
                old_total
            ) * 100

            if old_total != 0

            else 0
        )


    else:

        grand_total[col] = (
            comparison[col].sum()
        )


# =========================================================
# FINAL DISPLAY TABLE
# =========================================================

final_table = pd.concat(
    [
        comparison,

        pd.DataFrame(
            [grand_total]
        )
    ],
    ignore_index=True
)


# =========================================================
# KPI
# =========================================================

month1_total = comparison[
    f"{from_month_label} | Total"
].sum()


month2_total = comparison[
    f"{to_month_label} | Total"
].sum()


total_difference = (
    month2_total -
    month1_total
)


overall_growth = (

    (
        total_difference
        /
        month1_total
    ) * 100

    if month1_total != 0

    else 0
)


# =========================================================
# KPI CARDS
# =========================================================

k1, k2, k3, k4, k5 = st.columns(5)


with k1:

    st.metric(
        "🏢 Total Branches",
        len(comparison)
    )


with k2:

    st.metric(
        f"📅 {from_month_label}",
        f"{int(month1_total):,}"
    )


with k3:

    st.metric(
        f"📅 {to_month_label}",
        f"{int(month2_total):,}"
    )


with k4:

    st.metric(
        "↕ Difference",
        f"{int(total_difference):,}"
    )


with k5:

    st.metric(
        "📈 Growth",
        f"{overall_growth:.2f}%"
    )


# =========================================================
# MAIN TABLE
# =========================================================

st.subheader(
    "📋 Area-wise / Branch-wise Range Comparison"
)


try:

    # -----------------------------------------------------
    # Convert column names to strings
    # -----------------------------------------------------

    final_table.columns = (
        final_table.columns
        .astype(str)
    )


    # -----------------------------------------------------
    # Safe object columns
    # -----------------------------------------------------

    for col in final_table.columns:

        if final_table[col].dtype == "object":

            final_table[col] = (
                final_table[col]
                .fillna("")
                .astype(str)
            )


    st.dataframe(
        final_table,
        use_container_width=True,
        height=700
    )


except Exception:

    st.error(
        "Table display error occurred."
    )


    st.table(
        final_table
    )


# =========================================================
# RANGE EXPLANATION
# =========================================================

st.markdown("---")


c1, c2 = st.columns(2)


with c1:

    st.markdown(
        "### 📌 Recovery Ranges"
    )


    st.markdown("""
    **1–5** → Day 1 to Day 5  

    **6–10** → Day 6 to Day 10  

    **11–15** → Day 11 to Day 15  

    **16–25** → Day 16 to Day 25  

    **>25** → More than 25 Days
    """)


with c2:

    st.markdown(
        "### 📊 Comparison Formula"
    )


    st.info(
        f"""
        **Difference = {to_month_label} − {from_month_label}**

        **Growth % = Difference ÷ {from_month_label} × 100**

        ہر Area اور Branch کی recovery range-wise compare کی جا رہی ہے۔
        """
    )


# =========================================================
# DOWNLOAD EXCEL
# GREEN / RED DIFFERENCE REPORT
# =========================================================

# ---------------------------------------------------------
# CREATE WORKBOOK
# ---------------------------------------------------------

wb = Workbook()


ws = wb.active

ws.title = (
    "Range Wise Comparison"
)

ws.sheet_state = "visible"


# =========================================================
# WRITE HEADERS
# =========================================================

headers = list(
    final_table.columns
)


for col_num, header in enumerate(
    headers,
    start=1
):

    cell = ws.cell(
        row=1,
        column=col_num,
        value=header
    )


    cell.font = Font(
        bold=True,
        color="FFFFFF"
    )


    cell.fill = PatternFill(
        fill_type="solid",
        fgColor="063B66"
    )


    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


# =========================================================
# WRITE DATA
# =========================================================

for row_num, row_data in enumerate(

    final_table.itertuples(
        index=False
    ),

    start=2
):

    for col_num, value in enumerate(

        row_data,

        start=1

    ):

        cell = ws.cell(
            row=row_num,
            column=col_num,
            value=value
        )


        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


# =========================================================
# FIND DIFFERENCE COLUMNS
# =========================================================

difference_columns = []


for col_num in range(
    1,
    ws.max_column + 1
):

    header = ws.cell(
        row=1,
        column=col_num
    ).value


    if header:

        header_text = str(
            header
        )


        if (
            header_text.startswith(
                "Diff |"
            )
            or
            header_text == "% Diff"
        ):

            difference_columns.append(
                col_num
            )


# =========================================================
# GREEN / RED COLORS
# =========================================================

green_fill = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)


green_font = Font(
    color="006100",
    bold=True
)


red_fill = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)


red_font = Font(
    color="9C0006",
    bold=True
)


# =========================================================
# CONDITIONAL FORMATTING
# =========================================================

grand_total_row = ws.max_row

first_data_row = 2

last_data_row = (
    grand_total_row - 1
)


if last_data_row >= first_data_row:

    for col_num in difference_columns:

        col_letter = (
            get_column_letter(
                col_num
            )
        )


        data_range = (
            f"{col_letter}"
            f"{first_data_row}:"
            f"{col_letter}"
            f"{last_data_row}"
        )


        # -------------------------------------------------
        # POSITIVE = GREEN
        # -------------------------------------------------

        ws.conditional_formatting.add(

            data_range,

            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=green_fill,
                font=green_font
            )

        )


        # -------------------------------------------------
        # NEGATIVE = RED
        # -------------------------------------------------

        ws.conditional_formatting.add(

            data_range,

            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=red_fill,
                font=red_font
            )

        )


# =========================================================
# GRAND TOTAL ROW
# =========================================================

grand_fill = PatternFill(
    fill_type="solid",
    fgColor="063B66"
)


grand_font = Font(
    bold=True,
    color="FFFFFF"
)


for col_num in range(
    1,
    ws.max_column + 1
):

    cell = ws.cell(
        row=grand_total_row,
        column=col_num
    )


    cell.fill = grand_fill

    cell.font = grand_font


    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


# =========================================================
# % DIFFERENCE FORMAT
# =========================================================

for col_num in range(
    1,
    ws.max_column + 1
):

    header = ws.cell(
        row=1,
        column=col_num
    ).value


    if header == "% Diff":

        for row_num in range(
            2,
            ws.max_row + 1
        ):

            ws.cell(
                row=row_num,
                column=col_num
            ).number_format = (
                '0.00"%"'
            )


# =========================================================
# BORDERS
# =========================================================

thin_border = Border(

    left=Side(
        style="thin",
        color="D9E1F2"
    ),

    right=Side(
        style="thin",
        color="D9E1F2"
    ),

    top=Side(
        style="thin",
        color="D9E1F2"
    ),

    bottom=Side(
        style="thin",
        color="D9E1F2"
    )
)


for row in ws.iter_rows():

    for cell in row:

        cell.border = thin_border


# =========================================================
# COLUMN WIDTH
# =========================================================

for col_num in range(
    1,
    ws.max_column + 1
):

    max_length = 0


    for row_num in range(
        1,
        ws.max_row + 1
    ):

        value = ws.cell(
            row=row_num,
            column=col_num
        ).value


        if value is not None:

            max_length = max(
                max_length,
                len(str(value))
            )


    header = ws.cell(
        row=1,
        column=col_num
    ).value


    if header == "Branch Name":

        ws.column_dimensions[
            get_column_letter(
                col_num
            )
        ].width = 22


    elif header == "Area":

        ws.column_dimensions[
            get_column_letter(
                col_num
            )
        ].width = 20


    else:

        ws.column_dimensions[
            get_column_letter(
                col_num
            )
        ].width = min(
            max(
                max_length + 3,
                10
            ),
            18
        )


# =========================================================
# ROW HEIGHT
# =========================================================

ws.row_dimensions[1].height = 30


for row_num in range(
    2,
    ws.max_row + 1
):

    ws.row_dimensions[
        row_num
    ].height = 22


# =========================================================
# FREEZE PANES
# =========================================================

ws.freeze_panes = "D2"


# =========================================================
# AUTO FILTER
# =========================================================

ws.auto_filter.ref = (
    ws.dimensions
)


# =========================================================
# SAVE TO MEMORY
# =========================================================

excel_output = BytesIO()


wb.save(
    excel_output
)


excel_output.seek(0)


# =========================================================
# DOWNLOAD BUTTON
# =========================================================

st.download_button(

    label="⬇️ Download Range-Wise Excel Report",

    data=excel_output.getvalue(),

    file_name=(
        "Range_Wise_Recovery_Comparison.xlsx"
    ),

    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),

    use_container_width=True

)

import streamlit as st
import pandas as pd
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from zipfile import ZipFile

st.header("📑 Cheque-wise Analysis")

uploaded_cheque = st.file_uploader("Upload Cheque-wise List", type=["xlsx", "csv"])

if uploaded_cheque:

    if uploaded_cheque.name.endswith(".csv"):
        cheque_df = pd.read_csv(uploaded_cheque)
    else:
        cheque_df = pd.read_excel(uploaded_cheque)

    cheque_df.columns = [str(c).strip() for c in cheque_df.columns]

    required_cols = ["branch_id","date_disbursed","sanction_no","tranch_no","member_name","member_cnic"]
    cheque_df = cheque_df[[c for c in required_cols if c in cheque_df.columns]]

    cheque_df["Name"] = cheque_df["member_name"]
    cheque_df.drop(columns=["member_name"], inplace=True)

    cheque_df["date_disbursed"] = pd.to_datetime(cheque_df["date_disbursed"], errors="coerce")

    today = datetime.today()

    cheque_df["Months Passed"] = cheque_df["date_disbursed"].apply(
        lambda x: relativedelta(today, x).months + relativedelta(today, x).years*12 if pd.notnull(x) else ""
    )

    cheque_df["Days Passed"] = cheque_df["date_disbursed"].apply(
        lambda x: (today-x).days if pd.notnull(x) else ""
    )

    for col in ["House Complete","Shifted","Design"]:
        if col not in cheque_df.columns:
            cheque_df[col] = ""

    cheque_df["2nd Tranch Status"] = ""

    second_map = cheque_df[cheque_df["tranch_no"]==2].groupby("sanction_no").size().to_dict()

    first_df = cheque_df[cheque_df["tranch_no"]==1].copy()
    first_df["2nd Tranch Status"] = first_df["sanction_no"].apply(lambda x:"OK" if x in second_map else "")

    display_cols = ["branch_id","sanction_no","tranch_no","Name","member_cnic",
                    "date_disbursed","Months Passed","2nd Tranch Status",
                    "House Complete","Shifted","Design"]

    editable_df = first_df[display_cols]

    # -------- Editable Table --------
    edited_df = st.experimental_data_editor(editable_df, use_container_width=True)

    # -------- CSV Download --------
    csv_data = edited_df.to_csv(index=False).encode("utf-8")

    st.download_button(
        "⬇️ Download Edited CSV",
        csv_data,
        "cheque_analysis.csv",
        "text/csv"
    )

    # -------- Save Flags --------
    if st.button("💾 Save Flags"):
        edited_df[["sanction_no","tranch_no","House Complete","Shifted","Design"]].to_csv(
            "cheque_flags.csv", index=False
        )
        st.success("Saved")

    # -------- ZIP PDFs --------
    if st.button("⬇️ Download All Branch PDFs (ZIP)"):

        zip_buffer = BytesIO()

        with ZipFile(zip_buffer,"w") as zipf:

            for branch in edited_df["branch_id"].unique():

                bdf = edited_df[edited_df["branch_id"]==branch]

                pdf = BytesIO()
                doc = SimpleDocTemplate(pdf, pagesize=landscape(A4))
                styles = getSampleStyleSheet()
                elements=[]

                elements.append(Paragraph(f"Branch {branch}",styles["Heading1"]))
                elements.append(Spacer(1,10))

                table_df = bdf.drop(columns=["branch_id","tranch_no"],errors="ignore")
                table_df.insert(0,"S.No",range(1,len(table_df)+1))

                data=[table_df.columns.tolist()]+table_df.astype(str).values.tolist()

                table=Table(data,repeatRows=1)
                table.setStyle(TableStyle([
                    ("GRID",(0,0),(-1,-1),0.5,colors.black),
                    ("ALIGN",(0,0),(-1,-1),"CENTER")
                ]))

                elements.append(table)
                doc.build(elements)
                pdf.seek(0)

                zipf.writestr(f"branch_{branch}.pdf",pdf.getvalue())

        zip_buffer.seek(0)

        st.download_button(
            "Download ZIP",
            zip_buffer.getvalue(),
            "branches.zip",
            "application/zip"
        )
import streamlit as st
import pandas as pd
from fpdf import FPDF

st.title("Loan Disbursement PDF Generator (Branchwise)")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

# ---------------------- Safe Functions ----------------------
def safe(val):
    try:
        if pd.isna(val):
            return ""
        return str(val)
    except:
        return ""

# ---------------------- PDF Class ----------------------
class PDF(FPDF):
    def header(self):
        self.set_font("Arial", 'B', 12)
        self.cell(0, 8, "Loan Disbursement Report", ln=True, align="C")
        self.ln(3)

# ---------------------- PDF GENERATION FUNCTION ----------------------
# لاجک بالکل سیم ہے، بس فنکشن میں ڈالنے سے اسکرین پر None پرنٹ ہونا بند ہو جائے گا
def generate_pdf_data(br_name, br_data):
    pdf = PDF(orientation="L", unit="mm", format="A4")  # LANDSCAPE
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 8, f"Branch: {br_name}", ln=True, align="C")
    pdf.ln(3)

    # ---------------------- TABLE HEADER ----------------------
    headers = [
        "Date Disburse", "Sanction No", "Tranch", "Cheque No",
        "Loan Amount", "Group No", "Member Name", "CNIC"
    ]
    col_widths = [30, 35, 15, 40, 30, 30, 55, 45]

    pdf.set_fill_color(200, 200, 200)
    pdf.set_font("Arial", 'B', 9)

    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align="C", fill=True)
    pdf.ln()

    # ---------------------- TABLE ROWS ----------------------
    fill = False
    for _, row in br_data.iterrows():
        pdf.set_fill_color(235, 245, 255) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.set_font("Arial", '', 9)

        pdf.cell(col_widths[0], 7, safe(row["date_disburse"]), border=1, fill=True)
        pdf.cell(col_widths[1], 7, safe(row["sanction_no"]), border=1, fill=True)
        pdf.cell(col_widths[2], 7, safe(row["tranch"]), border=1, fill=True)
        pdf.cell(col_widths[3], 7, safe(row["cheque_no"]), border=1, fill=True)
        pdf.cell(col_widths[4], 7, safe(row["loan_amount"]), border=1, fill=True)
        pdf.cell(col_widths[5], 7, safe(row["group_no"]), border=1, fill=True)
        pdf.cell(col_widths[6], 7, safe(row["member_name"]), border=1, fill=True)
        pdf.cell(col_widths[7], 7, safe(row["member_cnic"]), border=1, fill=True)

        pdf.ln()
        fill = not fill

    return pdf.output(dest="S").encode("latin-1")

# ---------------------- MAIN ----------------------
if uploaded_file:
    df = pd.read_excel(uploaded_file)

    # Fix column spellings
    df.rename(columns={
        "date_disbursed": "date_disburse",
        "date_of_disbursement": "date_disburse",
        "tranch_no": "tranch",
        "grouo_no": "group_no",
    }, inplace=True)

    # Required Columns
    required_cols = [
        "branch_id", "member_name", "member_cnic", "loan_amount",
        "tranch", "cheque_no", "sanction_no",
        "group_no", "date_disburse"
    ]

    # Check Missing Columns
    missing = [c for c in required_cols if c not in df.columns]

    if missing:
        st.error(f"Missing columns: {missing}")
        st.stop()

    branches = df["branch_id"].unique()

    # ---------------------- MAIN LOOP ----------------------
    for br in branches:
        br_df = df[df["branch_id"] == br]

        st.markdown(f"### 📌 Branch: **{br}**")
        st.dataframe(br_df)

        # پی ڈی ایف ڈیٹا بیک گراؤنڈ میں بنے گا (کوئی None نہیں دکھے گا)
        pdf_bytes = generate_pdf_data(br, br_df)

        # سنگل کلک ڈاؤن لوڈ بٹن
        st.download_button(
            label=f"Download PDF for Branch {br}",
            data=pdf_bytes,
            file_name=f"{br}_Loan_Disbursement.pdf",
            mime="application/pdf",
            key=f"btn_{br}"  # Unique Key
        )
        
        st.write("---")

    st.success("All Branch PDF Buttons Ready!")
import streamlit as st
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import os

st.title("Recovery Date Range Summary")

# ---------------- Local storage folder ----------------
LOCAL_FILE = "data/recovery.xlsx"
os.makedirs("data", exist_ok=True)

# ---------------- File Upload ----------------
uploaded = st.file_uploader("Upload Recovery Excel / CSV", type=["xlsx", "csv"])

# --- If uploaded, save locally and store in session_state ---
if uploaded:
    if uploaded.name.endswith(".csv"):
        df = pd.read_csv(uploaded)
    else:
        df = pd.read_excel(uploaded)

    st.session_state["df"] = df
    df.to_excel(LOCAL_FILE, index=False)
    st.success("File uploaded and saved locally!")

# --- If no upload, check session_state or local file ---
elif "df" in st.session_state:
    df = st.session_state["df"]
    st.info("Using previously uploaded file from session.")
elif os.path.exists(LOCAL_FILE):
    df = pd.read_excel(LOCAL_FILE)
    st.session_state["df"] = df
    st.info("Loaded previously uploaded file from local storage.")
else:
    st.info("Please upload recovery file.")
    st.stop()

# ---------------- Column Selection ----------------
st.subheader("Available Columns")

date_col = st.selectbox("Select Date Column", df.columns)
branch_col = st.selectbox("Select Branch Column (branch_id)", df.columns)
area_col = None
if 'area_id' in df.columns:
    area_col = 'area_id'

# ---------------- Convert Date ----------------
df[date_col] = pd.to_datetime(
    df[date_col].astype(str).str.strip(),
    format="%Y-%b-%d",
    errors="coerce"
)
df = df.dropna(subset=[date_col, branch_col])
df["Day"] = df[date_col].dt.day
df = df[df["Day"].notna()]

# --- Updated Bins (1-5, 6-10, 11-15, 16-31) ---
df["Range"] = pd.cut(
    df["Day"],
    bins=[0, 5, 10, 15, 31],
    labels=["1-5", "6-10", "11-15", "16-31"]
)
if df["Range"].isna().all():
    st.error("Date column sahi format me nahi.")
    st.stop()

# ---------------- Pivot Table ----------------
pivot = pd.pivot_table(
    df,
    index=[branch_col],
    columns="Range",
    aggfunc="size",
    fill_value=0
)

# Ensure columns exist
for c in ["1-5", "6-10", "11-15", "16-31"]:
    if c not in pivot.columns:
        pivot[c] = 0

pivot["Total"] = pivot[["1-5", "6-10", "11-15", "16-31"]].sum(axis=1)

# Percentages
pivot["1-5 %"] = (pivot["1-5"] / pivot["Total"] * 100).round(2)
pivot["6-10 %"] = (pivot["6-10"] / pivot["Total"] * 100).round(2)
pivot["11-15 %"] = (pivot["11-15"] / pivot["Total"] * 100).round(2)
pivot["16-31 %"] = (pivot["16-31"] / pivot["Total"] * 100).round(2)

# Rename for readability
pivot.rename(columns={
    "1-5": "Recovery 1-5",
    "6-10": "Recovery 6-10",
    "11-15": "Recovery 11-15",
    "16-31": "Recovery 16-31"
}, inplace=True)

result_df = pivot.reset_index()

# ---------------- Add Area column BEFORE Branch ----------------
if area_col:
    branch_area_df = df[[branch_col, area_col]].drop_duplicates()
    result_df = result_df.merge(branch_area_df, on=branch_col, how='left')
    # Move Area column before Branch column
    cols = result_df.columns.tolist()
    branch_idx = cols.index(branch_col)
    cols.insert(branch_idx, cols.pop(cols.index(area_col)))
    result_df = result_df[cols]

# ---------------- Grand Total Row ----------------
numeric_cols = ["Recovery 1-5", "Recovery 6-10", "Recovery 11-15", "Recovery 16-31", "Total"]
# Sum numeric counts
grand_total_counts = result_df[numeric_cols].sum()
# Calculate percentages for Grand Total
grand_total_percent = (grand_total_counts[["Recovery 1-5", "Recovery 6-10", "Recovery 11-15", "Recovery 16-31"]] / grand_total_counts["Total"] * 100).round(2)

grand_values = {}
for col in result_df.columns:
    if col == branch_col:
        grand_values[col] = "Grand Total"
    elif col == area_col:
        grand_values[col] = ""
    elif col in numeric_cols:
        grand_values[col] = grand_total_counts[col]
    elif col in ["1-5 %", "6-10 %", "11-15 %", "16-31 %"]:
        pct_map = {
            "1-5 %": "Recovery 1-5", 
            "6-10 %": "Recovery 6-10", 
            "11-15 %": "Recovery 11-15", 
            "16-31 %": "Recovery 16-31"
        }
        grand_values[col] = grand_total_percent[pct_map[col]]
    else:
        grand_values[col] = ""

result_df = pd.concat([result_df, pd.DataFrame([grand_values])], ignore_index=True)

# ---------------- Show Table ----------------
st.subheader("Branch Wise Recovery Summary")
st.dataframe(result_df)

# ---------------- CSV Download ----------------
csv = result_df.to_csv(index=False).encode("utf-8")
st.download_button(
    label="⬇ Download CSV",
    data=csv,
    file_name="recovery_summary.csv",
    mime="text/csv"
)

# ---------------- PDF Download ----------------
buffer = BytesIO()
doc = SimpleDocTemplate(buffer, pagesize=A4)

# Table data
table_data = [result_df.columns.tolist()] + result_df.values.tolist()

# Create Table with style
table = Table(table_data)
style = TableStyle([
    ('GRID', (0,0), (-1,-1), 1, colors.black),
    ('BACKGROUND', (0,0), (-1,0), colors.grey),
    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ('FONTSIZE', (0,0), (-1,-1), 10),
    ('BOTTOMPADDING', (0,0), (-1,0), 6),
])
table.setStyle(style)

doc.build([table])
pdf_bytes = buffer.getvalue()
buffer.close()

st.download_button(
    label="⬇ Download PDF",
    data=pdf_bytes,
    file_name="recovery_summary.pdf",
    mime="application/pdf"
)

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# =========================================================
# PAGE CONFIG
# =========================================================

    page_title="MDP Comparison",
    page_icon="📊",
    layout="wide"
)


# =========================================================
# CSS
# =========================================================

st.markdown("""
<style>

.mdp-title {
    background: linear-gradient(135deg,#063b66,#0876b9);
    color:white;
    padding:18px;
    border-radius:12px;
    text-align:center;
    font-size:27px;
    font-weight:800;
    margin-bottom:20px;
}

</style>
""", unsafe_allow_html=True)


st.markdown(
    '<div class="mdp-title">📊 MDP MONTH-WISE COMPARISON REPORT</div>',
    unsafe_allow_html=True
)


# =========================================================
# UPLOAD
# =========================================================

st.subheader("📤 Upload MDP Excel")

uploaded_file = st.file_uploader(
    "Upload MDP Excel File",
    type=["xlsx", "xls"],
    key="mdp_comparison_upload"
)

if uploaded_file is None:

    st.info(
        "Please upload your MDP Excel file."
    )

    st.stop()


# =========================================================
# READ EXCEL
# =========================================================

try:

    excel_file = pd.ExcelFile(uploaded_file)

    sheet_name = st.selectbox(
        "Select Sheet",
        excel_file.sheet_names
    )

    raw_df = pd.read_excel(
        uploaded_file,
        sheet_name=sheet_name
    )

except Exception as e:

    st.error(
        f"Excel file read نہیں ہو سکی: {e}"
    )

    st.stop()


# =========================================================
# CLEAN COLUMN NAMES
# =========================================================

raw_df.columns = (
    raw_df.columns
    .astype(str)
    .str.strip()
    .str.lower()
    .str.replace(" ", "_", regex=False)
    .str.replace("-", "_", regex=False)
)


# =========================================================
# FIND COLUMN
# =========================================================

def find_column(df, possible_names):

    for col in possible_names:

        if col in df.columns:
            return col

    return None


# =========================================================
# AUTO DETECT COLUMNS
# =========================================================

area_col = find_column(
    raw_df,
    [
        "area",
        "area_name",
        "region",
        "zone"
    ]
)


branch_col = find_column(
    raw_df,
    [
        "branch_name",
        "branchname",
        "branch_title",
        "branch"
    ]
)


date_col = find_column(
    raw_df,
    [
        "date",
        "mdp_date",
        "recovery_date",
        "transaction_date",
        "disbursement_date"
    ]
)


amount_col = find_column(
    raw_df,
    [
        "amount",
        "mdp_amount",
        "total_amount",
        "given_amount",
        "paid_amount",
        "recovery_amount"
    ]
)


receipt_col = find_column(
    raw_df,
    [
        "receipts",
        "receipt",
        "receipt_count",
        "receipt_no",
        "slips",
        "slip_count",
        "total_receipts"
    ]
)


# =========================================================
# COLUMN MAPPING
# =========================================================

st.subheader("🔧 Column Mapping")

c1, c2, c3 = st.columns(3)


# ---------------------------------------------------------
# AREA
# ---------------------------------------------------------

with c1:

    if area_col is None:

        area_col = st.selectbox(
            "Area Column",
            ["None"] + list(raw_df.columns),
            key="mdp_area"
        )

        if area_col == "None":
            area_col = None

    else:

        st.success(
            f"Area: {area_col}"
        )


# ---------------------------------------------------------
# BRANCH
# ---------------------------------------------------------

with c2:

    if branch_col is None:

        branch_col = st.selectbox(
            "Branch Name Column",
            ["None"] + list(raw_df.columns),
            key="mdp_branch"
        )

        if branch_col == "None":
            branch_col = None

    else:

        st.success(
            f"Branch: {branch_col}"
        )


# ---------------------------------------------------------
# DATE
# ---------------------------------------------------------

with c3:

    if date_col is None:

        date_col = st.selectbox(
            "Month / Date Column",
            ["None"] + list(raw_df.columns),
            key="mdp_date"
        )

        if date_col == "None":
            date_col = None

    else:

        st.success(
            f"Date: {date_col}"
        )


c4, c5 = st.columns(2)


# ---------------------------------------------------------
# AMOUNT
# ---------------------------------------------------------

with c4:

    if amount_col is None:

        amount_col = st.selectbox(
            "Amount Column",
            ["None"] + list(raw_df.columns),
            key="mdp_amount"
        )

        if amount_col == "None":
            amount_col = None

    else:

        st.success(
            f"Amount: {amount_col}"
        )


# ---------------------------------------------------------
# RECEIPTS
# ---------------------------------------------------------

with c5:

    if receipt_col is None:

        receipt_col = st.selectbox(
            "Receipts Column",
            ["None"] + list(raw_df.columns),
            key="mdp_receipts"
        )

        if receipt_col == "None":
            receipt_col = None

    else:

        st.success(
            f"Receipts: {receipt_col}"
        )


# =========================================================
# REQUIRED CHECK
# =========================================================

missing = []

if area_col is None:
    missing.append("Area")

if branch_col is None:
    missing.append("Branch Name")

if date_col is None:
    missing.append("Date / Month")

if amount_col is None:
    missing.append("Amount")

if receipt_col is None:
    missing.append("Receipts")


if missing:

    st.error(
        "یہ columns ضروری ہیں: "
        + ", ".join(missing)
    )

    st.stop()


# =========================================================
# PREPARE DATA
# =========================================================

df = raw_df.copy()


df["_area"] = (
    df[area_col]
    .fillna("")
    .astype(str)
    .str.strip()
)


df["_branch"] = (
    df[branch_col]
    .fillna("")
    .astype(str)
    .str.strip()
)


# =========================================================
# DATE
# =========================================================

df["_date"] = pd.to_datetime(
    df[date_col],
    errors="coerce",
    dayfirst=True
)


invalid_dates = df["_date"].isna().sum()


if invalid_dates:

    st.warning(
        f"{invalid_dates} rows میں valid date نہیں ملی۔ "
        "یہ rows شامل نہیں ہوں گی۔"
    )


df = df[
    df["_date"].notna()
].copy()


if df.empty:

    st.error(
        "کوئی valid date نہیں ملی۔"
    )

    st.stop()


# =========================================================
# AMOUNT
# =========================================================

df["_amount"] = pd.to_numeric(
    df[amount_col],
    errors="coerce"
).fillna(0)


# =========================================================
# RECEIPTS
# =========================================================

df["_receipts"] = pd.to_numeric(
    df[receipt_col],
    errors="coerce"
)


# If receipt column is numeric
# use numeric values.
# Otherwise count rows.

if df["_receipts"].isna().all():

    df["_receipts"] = 1

else:

    df["_receipts"] = (
        df["_receipts"]
        .fillna(0)
    )


# =========================================================
# MONTH
# =========================================================

df["_month"] = (
    df["_date"]
    .dt.to_period("M")
)


available_months = sorted(
    df["_month"]
    .unique()
)


if not available_months:

    st.error(
        "کوئی month data نہیں ملا۔"
    )

    st.stop()


# =========================================================
# MONTH NAMES
# =========================================================

month_names = [
    x.strftime("%b-%y")
    for x in available_months
]


st.success(
    f"{len(month_names)} months found: "
    + ", ".join(month_names)
)


# =========================================================
# MONTH REPORT FUNCTION
# =========================================================

def create_month_report(data, month_period):

    temp = data[
        data["_month"] == month_period
    ].copy()


    if temp.empty:

        return pd.DataFrame()


    result = []


    grouped = (
        temp
        .groupby(
            [
                "_area",
                "_branch"
            ],
            dropna=False
        )
        .agg(
            Amount=(
                "_amount",
                "sum"
            ),

            Receipts=(
                "_receipts",
                "sum"
            )
        )
        .reset_index()
    )


    grouped = grouped.sort_values(
        [
            "_area",
            "_branch"
        ]
    )


    for _, row in grouped.iterrows():

        result.append({

            "Area": row["_area"],

            "Branch Name": row["_branch"],

            "Amount": float(
                row["Amount"]
            ),

            "Receipts": float(
                row["Receipts"]
            )

        })


    return pd.DataFrame(result)


# =========================================================
# CREATE ALL MONTH REPORTS
# =========================================================

monthly_reports = {}


for month_period in available_months:

    monthly_reports[
        month_period
    ] = create_month_report(
        df,
        month_period
    )


# =========================================================
# ALL BRANCHES
# =========================================================

all_branches = pd.concat(

    [
        report[
            [
                "Area",
                "Branch Name"
            ]
        ]

        for report in monthly_reports.values()

        if not report.empty
    ],

    ignore_index=True

).drop_duplicates()


if all_branches.empty:

    st.error(
        "Area / Branch data نہیں ملا۔"
    )

    st.stop()


all_branches = (
    all_branches
    .sort_values(
        [
            "Area",
            "Branch Name"
        ]
    )
    .reset_index(drop=True)
)


# =========================================================
# BUILD DISPLAY TABLE
# =========================================================

display_df = all_branches.copy()


# ---------------------------------------------------------
# Keep due values in session
# ---------------------------------------------------------

if "mdp_due_values" not in st.session_state:

    st.session_state.mdp_due_values = {}


# =========================================================
# ADD MONTH COLUMNS
# =========================================================

for month_period in available_months:

    label = month_period.strftime(
        "%b-%y"
    )

    report = monthly_reports[
        month_period
    ]


    if report.empty:

        temp = all_branches.copy()

        temp[
            "Amount"
        ] = 0

        temp[
            "Receipts"
        ] = 0

    else:

        temp = all_branches.merge(

            report,

            on=[
                "Area",
                "Branch Name"
            ],

            how="left"

        )

        temp[
            "Amount"
        ] = temp[
            "Amount"
        ].fillna(0)

        temp[
            "Receipts"
        ] = temp[
            "Receipts"
        ].fillna(0)


    # -----------------------------------------------------
    # Amount
    # -----------------------------------------------------

    display_df[
        f"{label} | Amount"
    ] = temp[
        "Amount"
    ].values


    # -----------------------------------------------------
    # Receipts
    # -----------------------------------------------------

    display_df[
        f"{label} | Receipts"
    ] = temp[
        "Receipts"
    ].values


    # -----------------------------------------------------
    # Due
    # -----------------------------------------------------

    due_values = []


    for i in range(
        len(display_df)
    ):

        key = (
            str(
                display_df.iloc[i]["Area"]
            ),
            str(
                display_df.iloc[i]["Branch Name"]
            ),
            str(label)
        )


        due_values.append(
            st.session_state.mdp_due_values.get(
                key,
                0.0
            )
        )


    display_df[
        f"{label} | Due"
    ] = due_values


    # -----------------------------------------------------
    # MDP PER BOX
    # Amount / Due
    # -----------------------------------------------------

    display_df[
        f"{label} | MDP Per Box"
    ] = np.where(

        display_df[
            f"{label} | Due"
        ] != 0,

        display_df[
            f"{label} | Amount"
        ]
        /
        display_df[
            f"{label} | Due"
        ],

        0

    )


    # -----------------------------------------------------
    # NOT GIVEN
    # Receipts - Due
    # -----------------------------------------------------

    display_df[
        f"{label} | Not Given"
    ] = (

        display_df[
            f"{label} | Receipts"
        ]

        -

        display_df[
            f"{label} | Due"
        ]

    )


# =========================================================
# SERIAL NUMBER
# =========================================================

display_df.insert(
    0,
    "Sr.",
    range(
        1,
        len(display_df) + 1
    )
)


# =========================================================
# MANUAL DUE ENTRY
# =========================================================

st.markdown("---")

st.subheader(
    "✏️ Enter Due Manually"
)

st.info(
    "صرف **Due** columns میں values enter کریں۔ "
    "MDP Per Box اور Not Given خود calculate ہوں گے۔"
)


# =========================================================
# EDITABLE DUE TABLE
# =========================================================

due_edit_df = display_df[
    [
        "Sr.",
        "Area",
        "Branch Name"
    ]
    +
    [
        f"{m.strftime('%b-%y')} | Due"
        for m in available_months
    ]
].copy()


# =========================================================
# MANUAL DUE ENTRY - STREAMLIT COMPATIBLE
# =========================================================

st.markdown("---")

st.subheader("✏️ Enter Due Manually")

st.info(
    "ہر Branch کے سامنے ہر Month کا Due enter کریں۔ "
    "MDP Per Box اور Not Given خود calculate ہوں گے۔"
)

edited_due = due_edit_df.copy()


# =========================================================
# HEADER
# =========================================================

header_cols = st.columns(
    3 + len(available_months)
)

with header_cols[0]:

    st.markdown("**Sr.**")

with header_cols[1]:

    st.markdown("**Area**")

with header_cols[2]:

    st.markdown("**Branch Name**")


for month_index, month_period in enumerate(
    available_months
):

    label = month_period.strftime("%b-%y")

    with header_cols[3 + month_index]:

        st.markdown(
            f"**{label} Due**"
        )


# =========================================================
# INPUT ROWS
# =========================================================

for i in range(len(edited_due)):

    row_cols = st.columns(
        3 + len(available_months)
    )


    # -----------------------------------------------------
    # SERIAL
    # -----------------------------------------------------

    with row_cols[0]:

        st.write(
            edited_due.loc[i, "Sr."]
        )


    # -----------------------------------------------------
    # AREA
    # -----------------------------------------------------

    with row_cols[1]:

        st.write(
            edited_due.loc[i, "Area"]
        )


    # -----------------------------------------------------
    # BRANCH
    # -----------------------------------------------------

    with row_cols[2]:

        st.write(
            edited_due.loc[i, "Branch Name"]
        )


    # -----------------------------------------------------
    # DUE INPUT
    # -----------------------------------------------------

    for month_index, month_period in enumerate(
        available_months
    ):

        label = month_period.strftime(
            "%b-%y"
        )

        due_col = (
            f"{label} | Due"
        )

        current_value = edited_due.loc[
            i,
            due_col
        ]

        if pd.isna(current_value):

            current_value = 0


        with row_cols[
            3 + month_index
        ]:

            new_value = st.number_input(
                f"Due {label}",
                min_value=0.0,
                value=float(
                    current_value
                ),
                step=1.0,
                key=f"mdp_due_{i}_{label}"
            )

            edited_due.loc[
                i,
                due_col
            ] = new_value


# =========================================================
# SAVE MANUAL DUE VALUES
# =========================================================

for i, row in edited_due.iterrows():

    area = str(
        row["Area"]
    )

    branch = str(
        row["Branch Name"]
    )


    for month_period in available_months:

        label = month_period.strftime(
            "%b-%y"
        )

        due_col = (
            f"{label} | Due"
        )

        value = row[due_col]


        if pd.isna(value):

            value = 0


        key = (
            area,
            branch,
            label
        )


        st.session_state.mdp_due_values[
            key
        ] = float(value)


# =========================================================
# REBUILD FINAL TABLE AFTER DUE ENTRY
# =========================================================

final_df = all_branches.copy()


# =========================================================
# MONTH DATA AGAIN
# =========================================================

for month_period in available_months:

    label = month_period.strftime(
        "%b-%y"
    )


    report = monthly_reports[
        month_period
    ]


    if report.empty:

        temp = all_branches.copy()

        temp["Amount"] = 0

        temp["Receipts"] = 0

    else:

        temp = all_branches.merge(

            report,

            on=[
                "Area",
                "Branch Name"
            ],

            how="left"

        )

        temp["Amount"] = (
            temp["Amount"]
            .fillna(0)
        )

        temp["Receipts"] = (
            temp["Receipts"]
            .fillna(0)
        )


    final_df[
        f"{label} | Amount"
    ] = temp[
        "Amount"
    ].values


    final_df[
        f"{label} | Receipts"
    ] = temp[
        "Receipts"
    ].values


    due_values = []


    for i in range(
        len(final_df)
    ):

        key = (

            str(
                final_df.iloc[i]["Area"]
            ),

            str(
                final_df.iloc[i]["Branch Name"]
            ),

            label

        )


        due_values.append(

            st.session_state.mdp_due_values.get(
                key,
                0.0
            )

        )


    final_df[
        f"{label} | Due"
    ] = due_values


    # -----------------------------------------------------
    # MDP PER BOX
    # Amount / Due
    # -----------------------------------------------------

    final_df[
        f"{label} | MDP Per Box"
    ] = np.where(

        final_df[
            f"{label} | Due"
        ] != 0,

        final_df[
            f"{label} | Amount"
        ]
        /
        final_df[
            f"{label} | Due"
        ],

        0

    )


    # -----------------------------------------------------
    # NOT GIVEN
    # Receipts - Due
    # -----------------------------------------------------

    final_df[
        f"{label} | Not Given"
    ] = (

        final_df[
            f"{label} | Receipts"
        ]

        -

        final_df[
            f"{label} | Due"
        ]

    )


# =========================================================
# DIFFERENCE COLUMNS
# =========================================================

# Difference is made between consecutive months.

for i in range(
    1,
    len(available_months)
):

    previous_month = (
        available_months[i - 1]
    )

    current_month = (
        available_months[i]
    )


    previous_label = (
        previous_month.strftime(
            "%b-%y"
        )
    )


    current_label = (
        current_month.strftime(
            "%b-%y"
        )
    )


    # -----------------------------------------------------
    # Amount Difference
    # -----------------------------------------------------

    final_df[
        f"Diff {current_label}-{previous_label} | Amount"
    ] = (

        final_df[
            f"{current_label} | Amount"
        ]

        -

        final_df[
            f"{previous_label} | Amount"
        ]

    )


    # -----------------------------------------------------
    # Receipts Difference
    # -----------------------------------------------------

    final_df[
        f"Diff {current_label}-{previous_label} | Receipts"
    ] = (

        final_df[
            f"{current_label} | Receipts"
        ]

        -

        final_df[
            f"{previous_label} | Receipts"
        ]

    )


    # -----------------------------------------------------
    # Due Difference
    # -----------------------------------------------------

    final_df[
        f"Diff {current_label}-{previous_label} | Due"
    ] = (

        final_df[
            f"{current_label} | Due"
        ]

        -

        final_df[
            f"{previous_label} | Due"
        ]

    )


    # -----------------------------------------------------
    # MDP PER BOX DIFFERENCE
    # -----------------------------------------------------

    final_df[
        f"Diff {current_label}-{previous_label} | MDP Per Box"
    ] = (

        final_df[
            f"{current_label} | MDP Per Box"
        ]

        -

        final_df[
            f"{previous_label} | MDP Per Box"
        ]

    )


    # -----------------------------------------------------
    # NOT GIVEN DIFFERENCE
    # -----------------------------------------------------

    final_df[
        f"Diff {current_label}-{previous_label} | Not Given"
    ] = (

        final_df[
            f"{current_label} | Not Given"
        ]

        -

        final_df[
            f"{previous_label} | Not Given"
        ]

    )


# =========================================================
# SERIAL
# =========================================================

final_df.insert(
    0,
    "Sr.",
    range(
        1,
        len(final_df) + 1
    )
)


# =========================================================
# GRAND TOTAL
# =========================================================

grand_total = {}


for col in final_df.columns:

    if col == "Sr.":

        grand_total[col] = ""


    elif col == "Area":

        grand_total[col] = ""


    elif col == "Branch Name":

        grand_total[col] = "GRAND TOTAL"


    elif "MDP Per Box" in col:

        # Weighted / overall MDP per box
        # = Total Amount / Total Due

        parts = col.split("|")

        if len(parts) == 2:

            month_label = (
                parts[0].strip()
            )

            amount_col_name = (
                f"{month_label} | Amount"
            )

            due_col_name = (
                f"{month_label} | Due"
            )

            total_amount = (
                final_df[
                    amount_col_name
                ].sum()
            )

            total_due = (
                final_df[
                    due_col_name
                ].sum()
            )

            grand_total[col] = (

                total_amount /
                total_due

                if total_due != 0

                else 0

            )

        else:

            grand_total[col] = (
                final_df[col].sum()
            )


    else:

        grand_total[col] = (
            final_df[col].sum()
        )


final_table = pd.concat(

    [
        final_df,

        pd.DataFrame(
            [grand_total]
        )

    ],

    ignore_index=True

)


# =========================================================
# DISPLAY
# =========================================================

st.markdown("---")

st.subheader(
    "📋 MDP Month-wise Comparison"
)


# =========================================================
# FORMAT DISPLAY COPY
# =========================================================

display_final = final_table.copy()


for col in display_final.columns:

    if col in [
        "Area",
        "Branch Name"
    ]:

        display_final[col] = (
            display_final[col]
            .astype(str)
        )

    elif col != "Sr.":

        display_final[col] = pd.to_numeric(
            display_final[col],
            errors="coerce"
        ).fillna(0)


# =========================================================
# SHOW TABLE
# =========================================================

st.dataframe(

    display_final,

    use_container_width=True,

    height=700

)


# =========================================================
# EXPLANATION
# =========================================================

st.markdown("---")

e1, e2 = st.columns(2)


with e1:

    st.markdown("### 📌 MDP Formula")

    st.info("""
**MDP Per Box = Amount ÷ Due**

مثال:

Amount = 6,000  
Due = 1,000  

**6,000 ÷ 1,000 = 6**

یعنی **MDP Per Box = 6**
""")


with e2:

    st.markdown("### 📌 Not Given")

    st.info("""
**Not Given = Receipts − Due**

مثال:

Receipts = 1,000  
Due = 500  

**1,000 − 500 = 500**

یعنی **500 Not Given**
""")


# =========================================================
# EXCEL DOWNLOAD
# =========================================================

st.markdown("---")

st.subheader(
    "📥 Download Excel"
)


wb = Workbook()

ws = wb.active

ws.title = "MDP Comparison"


# =========================================================
# HEADERS
# =========================================================

for col_num, header in enumerate(
    final_table.columns,
    start=1
):

    cell = ws.cell(
        row=1,
        column=col_num,
        value=header
    )

    cell.font = Font(
        bold=True,
        color="FFFFFF"
    )

    cell.fill = PatternFill(
        fill_type="solid",
        fgColor="063B66"
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


# =========================================================
# DATA
# =========================================================

for row_num, row_data in enumerate(

    final_table.itertuples(
        index=False
    ),

    start=2

):

    for col_num, value in enumerate(
        row_data,
        start=1
    ):

        cell = ws.cell(
            row=row_num,
            column=col_num,
            value=value
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


# =========================================================
# GRAND TOTAL STYLE
# =========================================================

grand_total_row = ws.max_row


grand_fill = PatternFill(
    fill_type="solid",
    fgColor="063B66"
)


grand_font = Font(
    bold=True,
    color="FFFFFF"
)


for col_num in range(
    1,
    ws.max_column + 1
):

    cell = ws.cell(
        row=grand_total_row,
        column=col_num
    )

    cell.fill = grand_fill

    cell.font = grand_font

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


# =========================================================
# GREEN / RED DIFFERENCE
# =========================================================

green_fill = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

green_font = Font(
    color="006100",
    bold=True
)


red_fill = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)

red_font = Font(
    color="9C0006",
    bold=True
)


for col_num in range(
    1,
    ws.max_column + 1
):

    header = ws.cell(
        row=1,
        column=col_num
    ).value


    if header and str(
        header
    ).startswith("Diff"):

        col_letter = get_column_letter(
            col_num
        )


        if grand_total_row > 2:

            data_range = (
                f"{col_letter}2:"
                f"{col_letter}"
                f"{grand_total_row - 1}"
            )


            ws.conditional_formatting.add(

                data_range,

                CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=green_fill,
                    font=green_font
                )

            )


            ws.conditional_formatting.add(

                data_range,

                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=red_fill,
                    font=red_font
                )

            )


# =========================================================
# NUMBER FORMAT
# =========================================================

for col_num in range(
    1,
    ws.max_column + 1
):

    header = ws.cell(
        row=1,
        column=col_num
    ).value


    if header and (
        "MDP Per Box" in str(header)
    ):

        for row_num in range(
            2,
            ws.max_row + 1
        ):

            ws.cell(
                row=row_num,
                column=col_num
            ).number_format = "0.00"


# =========================================================
# BORDERS
# =========================================================

thin_border = Border(

    left=Side(
        style="thin",
        color="D9E1F2"
    ),

    right=Side(
        style="thin",
        color="D9E1F2"
    ),

    top=Side(
        style="thin",
        color="D9E1F2"
    ),

    bottom=Side(
        style="thin",
        color="D9E1F2"
    )

)


for row in ws.iter_rows():

    for cell in row:

        cell.border = thin_border


# =========================================================
# COLUMN WIDTH
# =========================================================

for col_num in range(
    1,
    ws.max_column + 1
):

    header = ws.cell(
        row=1,
        column=col_num
    ).value


    if header == "Area":

        width = 20

    elif header == "Branch Name":

        width = 22

    else:

        max_length = 0

        for row_num in range(
            1,
            ws.max_row + 1
        ):

            value = ws.cell(
                row=row_num,
                column=col_num
            ).value


            if value is not None:

                max_length = max(
                    max_length,
                    len(str(value))
                )


        width = min(
            max(
                max_length + 3,
                12
            ),
            22
        )


    ws.column_dimensions[
        get_column_letter(
            col_num
        )
    ].width = width


# =========================================================
# ROW HEIGHT
# =========================================================

ws.row_dimensions[1].height = 30


for row_num in range(
    2,
    ws.max_row + 1
):

    ws.row_dimensions[
        row_num
    ].height = 22


# =========================================================
# FREEZE AREA ONLY
# =========================================================

ws.freeze_panes = "B2"


# =========================================================
# FILTER
# =========================================================

ws.auto_filter.ref = (
    ws.dimensions
)


# =========================================================
# SAVE
# =========================================================

excel_output = BytesIO()

wb.save(
    excel_output
)

excel_output.seek(0)


# =========================================================
# DOWNLOAD
# =========================================================

st.download_button(

    label="⬇️ Download MDP Comparison Excel",

    data=excel_output.getvalue(),

    file_name=(
        "MDP_Month_Wise_Comparison.xlsx"
    ),

    mime=(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),

    use_container_width=True

)
